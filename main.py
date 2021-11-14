from datetime import datetime
import os
import json

import openpyxl
from loguru import logger as log

import gsheets
import excel
import parser
import hash

LOYALITY_RATES = [
    (1000, 'Гость (Скидка 1%)'),
    (2000, 'Новичок (Скидка 2%)'),
    (3000, 'Кандидат (Скидка 3%)'),
    (4000, 'Персона (Скидка 4%)'),
    (5000, 'Солдат (Скидка 5%)'),
    (6000, 'Сержант (Скидка 6%)'),
    (7000, 'Лейтенант (Скидка 7%)'),
    (8000, 'Капитан (Скидка 8%)'),
    (9000, 'Майор (Скидка 9%)'),
    (12000, 'Подполковник (Скидка 10%)'),
    (15000, 'Полковник (Скидка 11%)'),
    (19000, 'Генерал (Скидка 12%)'),
    (24000, 'Герой (Скидка 13%)'),
    (30000, 'Космонавт (Скидка 14%)'),
    (35000, 'Маг (Скидка 15%)'),
    (45000, 'Магистр (Скидка 16%)'),
    (55000, 'Волшебник (Скидка 17%)'),
    (60000, 'Мастер (Скидка 18%)'),
    (75000, 'Грандмастер (Скидка 19%)'),
    (90000, 'Элита (Скидка 20%)'),
    (110000, 'Король (Скидка 21%)'),
    (150000, 'Император (Скидка 22%)'),
    (200000, 'Президент (Скидка 23%)'),
    (250000, 'Андроид (Скидка 24%)'),
    (300000, 'Киберспортсмен (Скидка 25%)'),
    (500000, 'Легенда (Скидка 30%)')
]

def loyality_rate(costs: float) -> str:
    get_status = 'Клиент (Скидки нет)'
    for get_rate, status in LOYALITY_RATES:
        if costs > get_rate:
            get_status = status
        else:
            break
    return get_status


if __name__ == '__main__':
    balance_bonus = 0.0
    base_table_key = '1EsL801iyUvi7TtcHOubsnKyYt37VgeCoVPDlcVNi2HA'
    guest_table_key = '1If8wlJl6UVvXbiUXoV39g_gbe7WVCJXRRZUk-qJ7X9c'
    upload_table_key = '1n5A-fkR0LYCdTbdiBVLcRZ4-AKjEoetCE2bgGJEIym0'
    checksum_file = 'checksum.lst'
    base_file_name = 'base.json'
    excel_file_name = 'base.xlsx'
    black_list_famaly = 'blacklist.lst'
    start_time = datetime.now()
    changes_trigger = True
    black_list_gen = True
    double_key_search = True
    gsheets_save = True

    log.add(
        "long.log",
        filter=lambda record: "long" in record["extra"], mode='w'
    )
    log.add(
        "short.log",
        filter=lambda record: "short" in record["extra"], mode='w'
    )
    log.add(
        "error.log",
        filter=lambda record: "wrong" in record["extra"], mode='w'
    )
    checksum_list = hash.checksum_list(checksum_file)

    path = os.path.join(os.getcwd(), 'bases')
    files = [
        os.path.join(path, get_file) 
        for get_file in os.listdir(path) 
        if (os.path.isfile(os.path.join(path, get_file))
                and get_file[0].isalpha() and 'xls' in get_file.split('.')[1])
    ]


    if os.path.exists(base_file_name):
        with open(base_file_name, 'r', encoding='utf-8') as base_data:
            numbers_base = json.load(base_data)
    else:
        changes_trigger = True

    gsheets_base = gsheets.gsheets_form_load(base_table_key)
    gsheets_base.update(
        gsheets.gsheets_guest_load(guest_table_key)
    )

    if checksum_list.get(base_table_key) and checksum_list[base_table_key] != hash.checksum_dict(gsheets_base):
        changes_trigger = True
    else:
        for file in files:
            if not hash.checksum_check(file, checksum_list):
                changes_trigger = True
                break

    if changes_trigger:
        numbers_base = gsheets_base
        checksum_list.update({base_table_key: hash.checksum_dict(gsheets_base)})
        for get_number, returned_dict in gsheets_base.items():
            if not numbers_base.get(get_number):
                numbers_base[get_number] = returned_dict
            else:
                numbers_base[get_number]['balance'] = max(
                    returned_dict['balance'],
                    numbers_base[get_number]['balance']
                )
                numbers_base[get_number]['total_costs'] = max(
                    returned_dict['total_costs'],
                    numbers_base[get_number]['total_costs']
                )
            numbers_base[get_number]['loyality'] = loyality_rate(
                numbers_base[get_number]['total_costs']
            )
        for file in files:
            wbook = openpyxl.load_workbook(file)
            sheet = wbook.active
            get_file_name = os.path.split(file)[1]

            for row in sheet.iter_rows(
                min_row=1, 
                max_row=sheet.max_row, 
                max_col=sheet.max_column
            ):

                get_row = [cell.value for cell in row]
                if get_row[3]:
                    cells = [
                        str(get_cell)
                        if get_cell else ''
                        for get_cell in get_row
                    ]

                    get_number, returned_dict = parser.data_exctraction(get_file_name, *cells)

                    if get_number:
                        if not numbers_base.get(get_number):
                            numbers_base[get_number] = returned_dict
                        else:
                            numbers_base[get_number]['balance'] = max(
                                returned_dict['balance'],
                                numbers_base[get_number]['balance']
                            )
                            numbers_base[get_number]['total_costs'] = max(
                                returned_dict['total_costs'],
                                numbers_base[get_number]['total_costs']
                            )
                            numbers_base[get_number]['file'].extend(returned_dict['file'])
                        numbers_base[get_number]['loyality'] = loyality_rate(
                            numbers_base[get_number]['total_costs']
                        )

            checksum_list.update({get_file_name: hash.checksum_gen(file)})

        with open(checksum_file, 'w', encoding='utf-8') as new_checksum_file:
            json.dump(checksum_list, new_checksum_file, indent=4)

        if balance_bonus:
            print(f'Add balance bonus {balance_bonus} RUB.')
            for number in numbers_base.keys():
                numbers_base[number]['balance'] += balance_bonus

        if excel_file_name:
            excel.data_save(numbers_base, excel_file_name)

        if gsheets_save:
            print('Saving to Google sheets...', end='')
            gsheets.gsheets_save(upload_table_key, numbers_base)
            print('OK')
            print('Saving json file...', end='')
            with open(base_file_name, 'w', encoding='utf-8') as json_file:
                json.dump(numbers_base, json_file, indent=4,)
            print('OK')

        if black_list_gen:
            print('Calculate blacklist...')
            blacklist = list()
            export_black_list = dict()
            with open(black_list_famaly, 'r', encoding='utf-8') as black_file:
                for name in black_file:
                    for key, value in numbers_base.items():
                        for keyword in value['name'].split():
                            if name.strip() == keyword:
                                export_black_list.update({key: value})
            print('Saving blacklist...', end='')
            gsheets.gsheets_save(upload_table_key, export_black_list, sheet=1)
            print('OK')

        if double_key_search:
            print('Calculate duplicates...')
            double_list = dict()
            counter = 0

            for key, value in numbers_base.items():
                for search_key, search_value in numbers_base.items():
                    if value['name'] == search_value['name'] and key != search_key:
                        matches = 0
                        for num_main, num_additional in zip(list(key[2:]), list(search_key[2:])):
                            if num_main == num_additional:
                                matches += 1
                        if matches >= 7:
                            double_list.update({key: value, search_key: search_value})

            print('Record double values...', end='')
            gsheets.gsheets_save(upload_table_key, double_list, sheet=2, debug=True)
            print('OK')
    else:
        print('Data files not changed. Skiped.')

    print(f'{len(numbers_base.keys())} records in base.')
    print(f"Time work: {datetime.now() - start_time}")
