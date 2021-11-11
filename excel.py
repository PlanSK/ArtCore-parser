import random
import openpyxl


def data_save(numbers_dict: dict, file_name: str) -> None:
    boder_style = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin'), 
        right=openpyxl.styles.borders.Side(style='thin'), 
        top=openpyxl.styles.borders.Side(style='thin'), 
        bottom=openpyxl.styles.borders.Side(style='thin')
    )
    writebook = openpyxl.Workbook()
    active_sheet = writebook.active
    active_sheet.title = 'User base'
    cell_names = [
        'N п/п',
        'Phone number',
        'User name',
        'Balance',
        'Total costs',
        'Loyality'
    ]
    
    dimensions = (
        ("A", 10),
        ("B", 15),
        ("C", 60),
        ("D", 15),
        ("E", 15),
        ("F", 20)
    )

    for cell, dim in dimensions:
        active_sheet.column_dimensions[cell].width = dim

    for cell, name in enumerate(cell_names, start=1):
        active_sheet.cell(row=1, column=cell).value = name

    for index, (get_number, value) in enumerate(numbers_dict.items(), start=1):
        value_list = [
            index,
            get_number,
            numbers_dict[get_number]['name'],
            numbers_dict[get_number]['balance'],
            numbers_dict[get_number]['total_costs'],
            numbers_dict[get_number]['loyality']
        ]
        for col_number, value in enumerate(value_list, start=1):
            active_sheet.cell(row=index + 1, column=col_number).value = value

        for column in range(4, 6):
            active_sheet.cell(row=index + 1, column=column).number_format = '0.00'

    for row in active_sheet.iter_rows(
            min_row=1,
            max_col=active_sheet.max_column,
            max_row=active_sheet.max_row):
        for cell in row:
            cell.border = boder_style

    print(f'Saving file... {file_name}')
    try:
        writebook.save(file_name)
    except PermissionError:
        writebook.save(file_name.split('.')[0]+str(
            random.randint(10000,99999))+'.xlsx'
        )