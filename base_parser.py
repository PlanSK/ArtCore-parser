import openpyxl
from loguru import logger as log
import random
import re
import os


def number_cell_processing(
    file: str, 
    get_str_number: str, 
    get_value_cell: str
) -> int:

    short = False
    long = False
    if re.findall(r'([3789]){1}\d{1,}', get_value_cell):
        for s in re.finditer(r'([3789]){1}\d{1,}', get_value_cell):
            get_match = s.group()
            if get_match[0] == '9' and len(get_match) == 10:
                return '+7' + get_match
            elif get_match[0] != '9' and len(get_match) == 11:
                return '+7' + get_match[1:]
            elif len(get_match) < 11:
                short = True
            elif len(get_match) > 11:
                long = True

        get_match = next(re.finditer(r'([3789]){1}\d{1,}', get_value_cell)).group()
        if short:
            log.bind(short=True).info(
                f"({file}) Short number value in {get_str_number} row. "
                f"Value ({len(get_match)}) '{get_value_cell}'"
            )
            return 0
        elif long:
            log.bind(long=True).info(
                f"({file}) Long number value in {get_str_number} row. "
                f"Value ({len(get_match)}) '{get_value_cell}'"
            )
            return 0
    else:
        log.bind(wrong=True).info(
            f"({file}) Wrong value in {get_str_number} row. "
            f"Value '{get_value_cell}'"
        )


def data_exctraction(
    file: str,
    order_number: str,
    _,
    personal_name: str,
    phone_number: str,
    get_balance: str,
    __,
    costs: str,
    *remaining
) -> dict():

    data_dict = dict()

    if order_number and phone_number:
        get_number = number_cell_processing(file, order_number, phone_number)
        if get_number:
            data_dict['name'] = re.sub(r'[^А-Яа-яёЁ\s]', '', personal_name).strip()

            balance = 0.0
            if get_row[4]:
                try: 
                    balance = float(get_balance.replace(',', '.'))
                except ValueError:
                    log.bind(wrong=True).error(
                        f"({file}) Error in balance in {order_number}. "
                        f"Value: {get_balance}. Skipped."
                    )
            data_dict['balance'] = balance

            total_costs = 0.0
            if get_row[6]:
                try:
                    total_costs = float(costs.split()[0].replace(',', '.'))
                except ValueError:
                    log.bind(wrong=True).error(
                        f"({file}) Error in total costs in {order_number}. "
                        f"Value: {costs}. Skipped."
                    )
            data_dict['total_costs'] = total_costs

            return get_number, data_dict

    return None, None


def data_save(numbers_dict: dict) -> None:
    boder_style = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(style='thin'), 
        right=openpyxl.styles.borders.Side(style='thin'), 
        top=openpyxl.styles.borders.Side(style='thin'), 
        bottom=openpyxl.styles.borders.Side(style='thin')
    )
    writebook = openpyxl.Workbook()
    active_sheet = writebook.active
    active_sheet.title = 'User base'
    cell_names = [
        'N п/п',
        'Phone number',
        'User name',
        'Balance',
        'Total costs',
    ]
    
    dimensions = (
        ("A", 10),
        ("B", 15),
        ("C", 60),
        ("D", 15),
        ("E", 15),
    )

    for cell, dim in dimensions:
        active_sheet.column_dimensions[cell].width = dim

    for cell, name in enumerate(cell_names, start=1):
        active_sheet.cell(row=1, column=cell).value = name

    for index, (get_number, value) in enumerate(numbers_dict.items(), start=1):
        value_list = [
            index,
            get_number,
            numbers_dict[get_number]['name'],
            numbers_dict[get_number]['balance'],
            numbers_dict[get_number]['total_costs']
        ]
        for col_number, value in enumerate(value_list, start=1):
            active_sheet.cell(row=index + 1, column=col_number).value = value

        for column in range(4, 6):
            active_sheet.cell(row=index + 1, column=column).number_format = '0.00'

    for row in active_sheet.iter_rows(
            min_row=1,
            max_col=active_sheet.max_column,
            max_row=active_sheet.max_row):
        for cell in row:
            cell.border = boder_style

    print('Saving file...')
    try:
        writebook.save('base.xlsx')
    except PermissionError:
        writebook.save('base'+str(random.randint(10000,99999))+'.xlsx')


if __name__ == '__main__':
    path = os.path.join(os.getcwd(), 'bases')
    files = [
        os.path.join(path, get_file) 
        for get_file in os.listdir(path) 
        if (os.path.isfile(os.path.join(path, get_file))
                and get_file[0].isalpha() and 'xls' in get_file.split('.')[1])
    ]
    
    log.add(
        "long.log",
        filter=lambda record: "long" in record["extra"], mode='w'
    )
    log.add(
        "short.log",
        filter=lambda record: "short" in record["extra"], mode='w'
    )
    log.add(
        "error.log",
        filter=lambda record: "wrong" in record["extra"], mode='w'
    )
    
    numbers_base = dict()

    for file in files:
        wbook = openpyxl.load_workbook(file)
        sheet = wbook.active
        print(f'Total records: {sheet.max_row}')

        for row in sheet.iter_rows(
            min_row=1, 
            max_row=sheet.max_row, 
            max_col=sheet.max_column
        ):

            get_row = [cell.value for cell in row]
            if get_row[3]:
                cells = [
                    str(get_cell)
                    if get_cell else ''
                    for get_cell in get_row
                ]

                get_number, returned_dict = data_exctraction(file, *cells)

                if get_number:
                    if not numbers_base.get(get_number):
                        numbers_base[get_number] = returned_dict
                    else:
                        numbers_base[get_number]['balance'] = max(
                            returned_dict['balance'],
                            numbers_base[get_number]['balance']
                        )
                        numbers_base[get_number]['total_costs'] = max(
                            returned_dict['total_costs'],
                            numbers_base[get_number]['total_costs']
                        )

    print(f'Total found: {len(numbers_base.keys())}')
    data_save(numbers_base)
